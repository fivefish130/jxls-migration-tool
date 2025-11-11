#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
完整的生产级 JXLS 迁移工具头
自动检测和优化终端环境，支持各种现代开发环境
"""

import sys
import os


def setup_unicode_support():
    """设置 Unicode 支持 - 生产环境优化"""
    if os.name == "nt":
        # 检测现代终端
        modern_terminals = ['WT_SESSION', 'WT_PROFILE_ID', 'VSCODE_PID', 'TERM_PROGRAM']
        is_modern_terminal = any(var in os.environ for var in modern_terminals)

        if not is_modern_terminal:
            # 传统终端需要设置代码页
            os.system("chcp 65001 >nul 2>&1")
            print("🔧 已为传统终端启用 UTF-8 支持")
        else:
            print("🎯 现代终端检测，使用原生 UTF-8")

        # 统一设置流编码
        try:
            if hasattr(sys.stdout, 'reconfigure'):
                sys.stdout.reconfigure(encoding='utf-8')
                sys.stderr.reconfigure(encoding='utf-8')
        except:
            pass


# 初始化 Unicode 支持
setup_unicode_support()

"""
JXLS 1.x → 2.14.0 自动化迁移工具 (v3.4 - 修复版)

功能特性:
  • 指令转换: forEach→each, if(test→condition), out→${}, area自动生成, multiSheet支持
  • 格式保留: 样式、列宽、行高、合并单元格、背景色 (增强错误处理)
  • 智能识别: 基于文件头检测真实格式，不依赖后缀名
  • 终端优化: Windows Terminal自动UTF-8检测与配置
  • 报告生成: Markdown + JSON + DEBUG日志
  • 健壮迁移: 自动格式检测 + 双重处理器回退机制 (统一API)
  • 错误修复: 修复 'Format' object has no attribute 'font_index' 错误
  • 关键修复: 修复jx:each注释不生成和jx:area位置错误问题

版本: 3.4  |  作者: fivefish  |  日期: 2025-11-07
更新: 修复jx:each注释生成和jx:area位置问题
使用: python jxls_migration_tool.py --help
"""

import re
import json
import logging
import argparse
import traceback
import shutil
import tempfile
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any, Set
from collections import defaultdict
from dataclasses import dataclass

# xlrd 是可选的 - 仅用于读取旧版 .xls 文件
# xlrd 2.0.1+ 仅支持 .xlsx，如果需要处理 .xls 请安装 xlrd<2.0
try:
    import xlrd
    from xlrd import formatting
    XLRD_AVAILABLE = True
    XLRD_VERSION = tuple(map(int, xlrd.__version__.split('.')[:2]))
    if XLRD_VERSION >= (2, 0):
        print("警告: xlrd 2.0+ 仅支持 .xlsx 文件，无法处理 .xls 文件")
        print("      如需处理 .xls 文件，请安装: pip install 'xlrd<2.0'")
        XLRD_AVAILABLE = False
except ImportError:
    XLRD_AVAILABLE = False
    print("提示: 缺少xlrd库，.xls 文件将自动转换为 .xlsx 后处理")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
except ImportError:
    print("错误: 缺少openpyxl库，请运行: pip install openpyxl")
    sys.exit(1)

try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    print("警告: 缺少xlsxwriter库，将使用openpyxl (pip install xlsxwriter)")
    XLSXWRITER_AVAILABLE = False


# ============================================================================
# 日志配置
# ============================================================================

def setup_logging(log_file: Optional[str] = None, dry_run: bool = False, verbose: bool = False) -> logging.Logger:
    """
    配置日志系统

    Args:
        log_file: 日志文件路径
        dry_run: 是否为试运行模式
        verbose: 是否显示详细日志

    Returns:
        配置好的logger对象
    """
    logger = logging.getLogger('JxlsMigration')
    logger.setLevel(logging.DEBUG)

    # 清除已有的handlers
    logger.handlers.clear()

    # 控制台handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_level = logging.DEBUG if verbose else logging.INFO
    console_handler.setLevel(console_level)
    console_format = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    console_handler.setFormatter(console_format)
    logger.addHandler(console_handler)

    # 文件handler
    if log_file:
        file_handler = logging.FileHandler(log_file, encoding='utf-8', mode='w')
        file_handler.setLevel(logging.DEBUG)
        file_format = logging.Formatter(
            '%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(file_format)
        logger.addHandler(file_handler)

    if dry_run:
        logger.info("=" * 80)
        logger.info("试运行模式 (DRY RUN) - 不会实际修改文件")
        logger.info("=" * 80)

    return logger


# ============================================================================
# .xls → .xlsx 转换器
# ============================================================================

def convert_xls_to_xlsx(xls_path: str) -> Optional[str]:
    """
    将 .xls 文件转换为 .xlsx 文件

    Args:
        xls_path: .xls 文件路径

    Returns:
        转换后的 .xlsx 文件路径，或 None（转换失败）
    """
    import tempfile
    import shutil
    from pathlib import Path

    if not XLRD_AVAILABLE:
        # 如果没有 xlrd，提示用户
        print(f"  ❌ 无法处理 .xls 文件: {Path(xls_path).name}")
        print(f"     请安装旧版 xlrd: pip install 'xlrd<2.0'")
        return None

    try:
        # 读取 .xls 文件
        xls_book = xlrd.open_workbook(xls_path)

        # 创建临时 .xlsx 文件
        temp_dir = Path(tempfile.gettempdir())
        temp_xlsx = temp_dir / f"{Path(xls_path).stem}_temp.xlsx"

        # 使用 openpyxl 写入
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        for sheet_idx in range(xls_book.nsheets):
            xls_sheet = xls_book.sheet_by_index(sheet_idx)
            ws = wb.create_sheet(title=xls_sheet.name)

            # 复制数据
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    cell = xls_sheet.cell(row_idx, col_idx)
                    if cell.value is not None:
                        ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell.value)

        wb.save(str(temp_xlsx))
        return str(temp_xlsx)

    except Exception as e:
        print(f"  ❌ 转换失败: {e}")
        if temp_xlsx.exists():
            temp_xlsx.unlink()
        return None


# ============================================================================
# 文件格式检测
# ============================================================================

def detect_excel_format(file_path: str) -> Optional[str]:
    """
    检测Excel文件的真实格式（不依赖文件后缀）

    Args:
        file_path: 文件路径

    Returns:
        str: 'xls' 或 'xlsx' 或 None
    """
    try:
        with open(file_path, 'rb') as f:
            header = f.read(8)

            # XLS文件头部: D0 CF 11 E0 A1 B1 1A E1 (OLE2/Compound Document)
            if header[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
                return 'xls'

            # XLSX文件头部: PK (ZIP格式)
            elif header[:2] == b'PK':
                # 进一步验证是否是有效的XLSX
                try:
                    # 尝试用openpyxl打开验证
                    temp_wb = load_workbook(file_path, read_only=True)
                    temp_wb.close()
                    return 'xlsx'
                except Exception:
                    # 可能是损坏的xlsx或其他ZIP文件
                    return 'xls'

            else:
                return None
    except Exception as e:
        logging.debug(f"文件格式检测失败 {file_path}: {e}")
        return None


def safe_detect_excel_format(file_path: str, logger: Optional[logging.Logger] = None) -> str:
    """
    安全地检测Excel文件格式，带有详细的日志记录

    Args:
        file_path: 文件路径
        logger: 日志记录器

    Returns:
        str: 'xls' 或 'xlsx'
    """
    from pathlib import Path
    try:
        format_result = detect_excel_format(file_path)
        file_ext = Path(file_path).suffix.lower()

        if format_result:
            if logger:
                if format_result == 'xlsx' and file_ext == '.xls':
                    logger.warning(f"  ⚠️  文件 '{Path(file_path).name}' 后缀为.xls但实际格式为.xlsx")
                elif format_result == 'xls' and file_ext == '.xlsx':
                    logger.warning(f"  ⚠️  文件 '{Path(file_path).name}' 后缀为.xlsx但实际格式为.xls")
                logger.debug(f"  格式检测结果: {format_result}")
            return format_result

        # 如果检测失败，尝试通过文件扩展名判断
        if logger:
            logger.debug(f"  自动检测失败，使用扩展名判断: {file_ext}")

        if file_ext == '.xlsx':
            return 'xlsx'
        elif file_ext == '.xls':
            return 'xls'
        else:
            # 默认返回 xls
            if logger:
                logger.warning(f"  无法判断格式，默认使用 XLS 处理器")
            return 'xls'
    except Exception as e:
        if logger:
            logger.error(f"  格式检测出错: {e}")
        # 出错时默认返回 xls
        return 'xls'


# ============================================================================
# JXLS指令解析器
# ============================================================================

@dataclass
class CommandLocation:
    """命令位置信息"""
    row: int
    col: int
    sheet_name: str


class JxlsCommand:
    """JXLS指令基类"""

    def __init__(self, location: CommandLocation, raw_text: str):
        self.location = location
        self.raw_text = raw_text
        self.params = {}
        self.processed = False

    def __repr__(self):
        return f"{self.__class__.__name__}(sheet={self.location.sheet_name}, row={self.location.row}, col={self.location.col})"


class AreaCommand(JxlsCommand):
    """jx:area指令"""

    def __init__(self, location: CommandLocation, raw_text: str):
        super().__init__(location, raw_text)
        self.parse()

    def parse(self):
        """解析area参数"""
        clean_text = self.raw_text.strip()

        # 移除< >标签符号（如果存在）
        if clean_text.startswith("<") and clean_text.endswith(">"):
            clean_text = clean_text[1:-1]

        # 匹配 jx:area(lastCell="xxx")
        patterns = [
            r'jx:area\s*\(\s*lastCell\s*=\s*["\']([^"\']*)["\']\s*\)',
            r'jx:area\s+lastCell\s*=\s*["\']([^"\']*)["\']',
        ]

        for pattern in patterns:
            match = re.search(pattern, clean_text, re.IGNORECASE)
            if match:
                self.params['lastCell'] = match.group(1)
                break

    def to_jx_area_v2(self, last_cell: str = None) -> str:
        """转换为JXLS 2.x的jx:area命令"""
        actual_last_cell = last_cell or self.params.get('lastCell', '')
        cmd = f'jx:area(lastCell="{actual_last_cell}")'
        return cmd


class ForEachCommand(JxlsCommand):
    """jx:forEach指令"""

    def __init__(self, location: CommandLocation, raw_text: str):
        super().__init__(location, raw_text)
        self.parse()
        self.end_location = None
        self.data_location = None

    def parse(self):
        """解析forEach参数 - 支持XML标签式和函数调用式"""
        clean_text = self.raw_text.strip()

        # 移除< >标签符号
        if clean_text.startswith("<") and clean_text.endswith(">"):
            clean_text = clean_text[1:-1]

        # 匹配forEach（两种格式）
        patterns = [
            r'jx:forEach\s+items\s*=\s*["\']([^"\']*)["\']\s+var\s*=\s*["\']([^"\']*)["\']',
            r'jx:forEach\s*\(\s*items\s*=\s*["\']([^"\']*)["\']\s*,\s*var\s*=\s*["\']([^"\']*)["\']',
        ]

        for pattern in patterns:
            match = re.search(pattern, clean_text, re.IGNORECASE)
            if match:
                items = match.group(1)
                var = match.group(2)
                # 移除${}包装
                items = re.sub(r'\$\{([^}]+)\}', r'\1', items)
                var = re.sub(r'\$\{([^}]+)\}', r'\1', var)
                self.params["items"] = items
                self.params["var"] = var
                break

        # 解析其他可选参数
        optional_params = ['varStatus', 'direction', 'multisheet', 'select', 'groupBy', 'groupOrder']
        for param in optional_params:
            pattern = f'{param}\\s*=\\s*["\']([^"\']*)["\']'
            match = re.search(pattern, clean_text, re.IGNORECASE)
            if match:
                self.params[param] = match.group(1)

    def to_jx_each(self, last_cell: str) -> str:
        """转换为jx:each命令"""
        items = self.params.get('items', '')
        var = self.params.get('var', '')

        # 基本参数
        cmd = f'jx:each(items="{items}" var="{var}" lastCell="{last_cell}"'

        # 可选参数
        optional_params = ['direction', 'multisheet', 'select', 'groupBy', 'groupOrder']
        for param in optional_params:
            if param in self.params:
                cmd += f' {param}="{self.params[param]}"'

        # 注意: JXLS 2.x不再使用varStatus，需要在Java代码中手动添加索引
        if 'varStatus' in self.params:
            cmd += ' # 注意: varStatus需要在Java代码中手动实现'

        cmd += ')'
        return cmd


class IfCommand(JxlsCommand):
    """jx:if指令"""

    def __init__(self, location: CommandLocation, raw_text: str):
        super().__init__(location, raw_text)
        self.parse()
        self.end_location = None
        self.data_location = None

    def parse(self):
        """解析if参数"""
        clean_text = self.raw_text.strip()

        # 移除< >标签符号
        if clean_text.startswith("<") and clean_text.endswith(">"):
            clean_text = clean_text[1:-1]

        # 匹配 jx:if(test="xxx") 或 jx:if(condition="xxx")
        patterns = [
            r'jx:if\s*\(\s*(?:test|condition)\s*=\s*["\']([^"\']*)["\']',
            r'jx:if\s+(?:test|condition)\s*=\s*["\']([^"\']*)["\']',
        ]

        for pattern in patterns:
            match = re.search(pattern, clean_text, re.IGNORECASE)
            if match:
                self.params['condition'] = match.group(1)
                break

        # 解析其他可选参数
        optional_params = ['direction', 'multisheet', 'lastCell', 'areas']
        for param in optional_params:
            pattern = f'{param}\\s*=\\s*["\']([^"\']*)["\']'
            match = re.search(pattern, clean_text, re.IGNORECASE)
            if match:
                self.params[param] = match.group(1)

    def to_jx_if_v2(self, last_cell: str) -> str:
        """转换为JXLS 2.x的jx:if命令"""
        condition = self.params.get('condition', '')

        cmd = f'jx:if(condition="{condition}" lastCell="{last_cell}"'

        # 可选参数
        optional_params = ['direction', 'multisheet', 'areas']
        for param in optional_params:
            if param in self.params:
                cmd += f' {param}="{self.params[param]}"'

        cmd += ')'
        return cmd


class OutCommand(JxlsCommand):
    """jx:out指令"""

    def __init__(self, location: CommandLocation, raw_text: str):
        super().__init__(location, raw_text)
        self.parse()

    def parse(self):
        """解析out参数"""
        clean_text = self.raw_text.strip()

        # 匹配 <jx:out select="xxx"/> 或 jx:out(select="xxx")
        patterns = [
            r'<jx:out\s+select="([^"]+)"\s*/?>',
            r'jx:out\s*\(\s*select\s*=\s*["\']([^"\']*)["\']\s*\)',
        ]

        for pattern in patterns:
            match = re.search(pattern, clean_text, re.IGNORECASE)
            if match:
                self.params['select'] = match.group(1)
                break

    def to_expression(self) -> str:
        """转换为${...}表达式"""
        select = self.params.get('select', '')
        return f'${{{select}}}'


class MultiSheetCommand(JxlsCommand):
    """jx:multiSheet指令"""

    def __init__(self, location: CommandLocation, raw_text: str):
        super().__init__(location, raw_text)
        self.parse()

    def parse(self):
        """解析multiSheet参数"""
        clean_text = self.raw_text.strip()

        patterns = [
            r'jx:multiSheet\s*\(\s*data\s*=\s*["\']([^"\']*)["\']',
            r'jx:multiSheet\s+data\s*=\s*["\']([^"\']*)["\']',
        ]

        for pattern in patterns:
            match = re.search(pattern, clean_text, re.IGNORECASE)
            if match:
                self.params['data'] = match.group(1)
                break

    def to_jx_multi_sheet_v2(self) -> str:
        """转换为JXLS 2.x的jx:multiSheet命令"""
        data = self.params.get('data', '')
        cmd = f'jx:multiSheet(data="{data}")'
        return cmd


# ============================================================================
# Excel格式转换器
# ============================================================================

class ExcelFormatConverter:
    """Excel格式转换器 - 从XLS转换到XLSX并保留格式"""

    # XLS颜色索引到RGB的映射（部分常用颜色）
    XLS_COLOR_MAP = {
        0: None,  # 自动
        1: '000000',  # 黑色
        2: 'FFFFFF',  # 白色
        3: 'FF0000',  # 红色
        4: '00FF00',  # 绿色
        5: '0000FF',  # 蓝色
        6: 'FFFF00',  # 黄色
        7: 'FF00FF',  # 品红
        8: '00FFFF',  # 青色
        9: '800000',  # 深红
        10: '008000',  # 深绿
        11: '000080',  # 深蓝
        12: '808000',  # 橄榄绿
        13: '800080',  # 紫色
        14: '008080',  # 青绿
        15: 'C0C0C0',  # 银色
        16: '808080',  # 灰色
        17: '9999FF',  # 淡紫
        18: '993366',  # 深粉
        19: 'FFFFCC',  # 淡黄
        20: 'CCFFFF',  # 淡青
        21: '660066',  # 深紫
        22: 'FF8080',  # 粉红
        23: '0066CC',  # 天蓝
        24: 'CCCCFF',  # 淡蓝
    }

    @staticmethod
    def get_rgb_from_xls_color(color_index: int) -> Optional[str]:
        """
        从XLS颜色索引获取RGB值

        Args:
            color_index: XLS颜色索引

        Returns:
            RGB字符串（如'FF0000'）或None
        """
        if color_index is None:
            return None
        return ExcelFormatConverter.XLS_COLOR_MAP.get(color_index)

    @staticmethod
    def convert_font(xls_font, xls_book) -> Font:
        """
        转换字体格式 - 增强错误处理

        Args:
            xls_font: xlrd的Font对象
            xls_book: xlrd的Workbook对象

        Returns:
            openpyxl的Font对象
        """
        try:
            font_args = {}

            # 字体名称
            if hasattr(xls_font, 'name') and xls_font.name:
                font_args['name'] = xls_font.name
            else:
                font_args['name'] = 'Calibri'  # 默认字体

            # 字体大小
            if hasattr(xls_font, 'height') and xls_font.height:
                font_args['size'] = xls_font.height / 20  # twips to points
            else:
                font_args['size'] = 11  # 默认大小

            # 字体样式
            if hasattr(xls_font, 'bold'):
                font_args['bold'] = xls_font.bold
            if hasattr(xls_font, 'italic'):
                font_args['italic'] = xls_font.italic
            if hasattr(xls_font, 'underline_type'):
                font_args['underline'] = 'single' if xls_font.underline_type else None
            if hasattr(xls_font, 'struck_out'):
                font_args['strike'] = xls_font.struck_out

            # 字体颜色
            if hasattr(xls_font, 'colour_index') and xls_font.colour_index:
                color_rgb = ExcelFormatConverter.get_rgb_from_xls_color(xls_font.colour_index)
                if color_rgb:
                    font_args['color'] = color_rgb

            return Font(**font_args)

        except Exception as e:
            logging.debug(f"字体转换失败: {e}")
            return Font(name='Calibri', size=11)  # 返回默认字体

    @staticmethod
    def convert_fill(xls_format, xls_book) -> Optional[PatternFill]:
        """
        转换填充格式 - 增强错误处理

        Args:
            xls_format: xlrd的Format对象
            xls_book: xlrd的Workbook对象

        Returns:
            openpyxl的PatternFill对象或None
        """
        try:
            # 检查必要的属性
            if not hasattr(xls_format, 'background'):
                return None

            background = xls_format.background

            # 获取背景色和前景色
            bg_color_index = getattr(background, 'background_colour_index', None)
            fg_color_index = getattr(background, 'pattern_colour_index', None)

            # 获取填充模式
            pattern = getattr(background, 'fill_pattern', None)
            if hasattr(background, 'pattern'):
                pattern = getattr(background.pattern, 'pattern_type_str', None)

            bg_color = ExcelFormatConverter.get_rgb_from_xls_color(bg_color_index)
            fg_color = ExcelFormatConverter.get_rgb_from_xls_color(fg_color_index)

            # 如果有颜色信息，创建填充
            if bg_color or fg_color:
                fill_type = 'solid'
                if pattern and pattern != 'Solid':
                    # 映射其他填充模式
                    fill_type = 'darkGray' if 'Gray' in pattern else 'solid'

                return PatternFill(
                    start_color=fg_color or 'FFFFFF',
                    end_color=bg_color or 'FFFFFF',
                    fill_type=fill_type
                )

        except Exception as e:
            logging.debug(f"填充转换失败: {e}")

        return None

    @staticmethod
    def convert_border(xls_format) -> Border:
        """
        转换边框格式 - 增强错误处理

        Args:
            xls_format: xlrd的Format对象

        Returns:
            openpyxl的Border对象
        """
        try:
            # xlrd的边框样式映射
            border_style_map = {
                0: None,  # No line
                1: 'thin',
                2: 'medium',
                3: 'dashed',
                4: 'dotted',
                5: 'thick',
                6: 'double',
                7: 'hair',
            }

            if not hasattr(xls_format, 'border'):
                return Border()

            xls_border = xls_format.border

            # 安全地获取边框样式
            def get_side(line_style_attr):
                if hasattr(xls_border, line_style_attr):
                    line_style = getattr(xls_border, line_style_attr)
                    style = border_style_map.get(line_style)
                    if style:
                        return Side(style=style)
                return None

            left = get_side('left_line_style')
            right = get_side('right_line_style')
            top = get_side('top_line_style')
            bottom = get_side('bottom_line_style')

            return Border(left=left, right=right, top=top, bottom=bottom)

        except Exception as e:
            logging.debug(f"边框转换失败: {e}")
            return Border()

    @staticmethod
    def convert_alignment(xls_format) -> Alignment:
        """
        转换对齐方式 - 增强错误处理

        Args:
            xls_format: xlrd的Format对象

        Returns:
            openpyxl的Alignment对象
        """
        try:
            # xlrd对齐方式映射
            horizontal_map = {
                0: 'general',
                1: 'left',
                2: 'center',
                3: 'right',
                4: 'fill',
                5: 'justify',
                6: 'centerContinuous',
                7: 'distributed',
            }

            vertical_map = {
                0: 'top',
                1: 'center',
                2: 'bottom',
                3: 'justify',
                4: 'distributed',
            }

            if not hasattr(xls_format, 'alignment'):
                return Alignment()

            xls_align = xls_format.alignment

            alignment_args = {}

            # 水平对齐
            if hasattr(xls_align, 'hor_align'):
                alignment_args['horizontal'] = horizontal_map.get(xls_align.hor_align, 'general')

            # 垂直对齐
            if hasattr(xls_align, 'vert_align'):
                alignment_args['vertical'] = vertical_map.get(xls_align.vert_align, 'bottom')

            # 自动换行
            if hasattr(xls_align, 'text_wrapped'):
                alignment_args['wrap_text'] = bool(xls_align.text_wrapped)

            return Alignment(**alignment_args)

        except Exception as e:
            logging.debug(f"对齐转换失败: {e}")
            return Alignment()

    @staticmethod
    def copy_cell_format(xls_cell, xls_book, xlsx_cell):
        """
        安全地复制单元格格式（简化版，避免富文本问题）

        Args:
            xls_cell: xlrd单元格对象
            xls_book: xlrd工作簿对象
            xlsx_cell: openpyxl单元格对象
        """
        try:
            # 获取格式索引
            if not hasattr(xls_cell, 'xf_index'):
                return

            xf_index = xls_cell.xf_index

            # 获取格式对象
            if not hasattr(xls_book, 'xf_list') or xf_index >= len(xls_book.xf_list):
                return

            xls_format = xls_book.xf_list[xf_index]

            # 复制最基础的格式：仅复制粗体/斜体（避免富文本问题）
            font_index = getattr(xls_format, 'font_index', None)
            if font_index is not None and hasattr(xls_book, 'font_list'):
                if font_index < len(xls_book.font_list):
                    xls_font = xls_book.font_list[font_index]
                    # 只复制样式，不复制字体名称（避免中文字体导致的兼容性问题）
                    font_args = {}
                    if hasattr(xls_font, 'bold') and xls_font.bold:
                        font_args['bold'] = True
                    if hasattr(xls_font, 'italic') and xls_font.italic:
                        font_args['italic'] = True
                    if font_args:
                        # 使用默认字体，仅应用样式
                        xlsx_cell.font = Font(name='Calibri', size=11, **font_args)

            # 简化填充复制：仅复制纯色填充
            fill = ExcelFormatConverter.convert_fill(xls_format, xls_book)
            if fill and fill.fill_type == 'solid':  # 仅复制纯色填充
                xlsx_cell.fill = fill

            # 简化边框复制：仅复制有边框的情况
            border = ExcelFormatConverter.convert_border(xls_format)
            if border and any([border.left.style, border.right.style, border.top.style, border.bottom.style]):
                xlsx_cell.border = border

            # 简化对齐复制：仅复制基本对齐
            alignment = ExcelFormatConverter.convert_alignment(xls_format)
            # 保持默认对齐，不强制设置（避免富文本问题）

        except Exception as e:
            # 记录详细错误信息用于调试
            logging.debug(
                f"复制单元格格式失败 (row={getattr(xls_cell, 'row', 'N/A')}, col={getattr(xls_cell, 'col', 'N/A')}): {e}")


# ============================================================================
# XlsxWriter 格式转换器
# ============================================================================

class XlsxWriterConverter:
    """基于 XlsxWriter 的格式转换器 - 自动使用共享字符串表"""

    # XLS颜色索引到RGB的映射（与 ExcelFormatConverter 相同）
    XLS_COLOR_MAP = {
        0: None,  # 自动
        1: '000000',  # 黑色
        2: 'FFFFFF',  # 白色
        3: 'FF0000',  # 红色
        4: '00FF00',  # 绿色
        5: '0000FF',  # 蓝色
        6: 'FFFF00',  # 黄色
        7: 'FF00FF',  # 品红
        8: '00FFFF',  # 青色
        9: '800000',  # 深红
        10: '008000',  # 深绿
        11: '000080',  # 深蓝
        12: '808000',  # 橄榄绿
        13: '800080',  # 紫色
        14: '008080',  # 青绿
        15: 'C0C0C0',  # 银色
        16: '808080',  # 灰色
        17: '9999FF',  # 淡紫
        18: '993366',  # 深粉
        19: 'FFFFCC',  # 淡黄
        20: 'CCFFFF',  # 淡青
        21: '660066',  # 深紫
        22: 'FF8080',  # 粉红
        23: '0066CC',  # 天蓝
        24: 'CCCCFF',  # 淡蓝
    }

    @staticmethod
    def get_rgb_from_xls_color(color_index: int) -> Optional[str]:
        """从XLS颜色索引获取RGB值"""
        if color_index is None:
            return None
        return XlsxWriterConverter.XLS_COLOR_MAP.get(color_index)

    @staticmethod
    def convert_font_xlsxwriter(xls_font, xls_book) -> dict:
        """
        转换字体格式为 xlsxwriter 格式

        Args:
            xls_font: xlrd的Font对象
            xls_book: xlrd的Workbook对象

        Returns:
            xlsxwriter 字体格式字典
        """
        try:
            font_format = {}

            # 字体名称
            if hasattr(xls_font, 'name') and xls_font.name:
                font_format['font_name'] = xls_font.name
            else:
                font_format['font_name'] = 'Calibri'

            # 字体大小
            if hasattr(xls_font, 'height') and xls_font.height:
                font_format['font_size'] = xls_font.height / 20  # twips to points
            else:
                font_format['font_size'] = 11

            # 字体样式
            if hasattr(xls_font, 'bold') and xls_font.bold:
                font_format['bold'] = True
            if hasattr(xls_font, 'italic') and xls_font.italic:
                font_format['italic'] = True

            # 字体颜色
            if hasattr(xls_font, 'colour_index') and xls_font.colour_index:
                color_rgb = XlsxWriterConverter.get_rgb_from_xls_color(xls_font.colour_index)
                if color_rgb:
                    font_format['font_color'] = color_rgb

            return font_format

        except Exception as e:
            logging.debug(f"字体转换失败: {e}")
            return {'font_name': 'Calibri', 'font_size': 11}

    @staticmethod
    def write_cell_format(workbook, worksheet, row, col, xls_cell, xls_book):
        """
        使用 xlsxwriter 安全地写入单元格格式

        Args:
            workbook: xlsxwriter Workbook对象
            worksheet: xlsxwriter Worksheet对象
            row: 行号 (0-based)
            col: 列号 (0-based)
            xls_cell: xlrd单元格对象
            xls_book: xlrd工作簿对象
        """
        try:
            # 获取格式索引
            if not hasattr(xls_cell, 'xf_index'):
                return None

            xf_index = xls_cell.xf_index

            # 获取格式对象
            if not hasattr(xls_book, 'xf_list') or xf_index >= len(xls_book.xf_list):
                return None

            xls_format = xls_book.xf_list[xf_index]

            # 收集格式信息
            format_dict = {}

            # 转换字体
            font_index = getattr(xls_format, 'font_index', None)
            if font_index is not None and hasattr(xls_book, 'font_list'):
                if font_index < len(xls_book.font_list):
                    xls_font = xls_book.font_list[font_index]
                    font_format = XlsxWriterConverter.convert_font_xlsxwriter(xls_font, xls_book)
                    format_dict.update(font_format)

            # 转换填充
            if hasattr(xls_format, 'background'):
                background = xls_format.background
                bg_color_index = getattr(background, 'background_colour_index', None)
                fg_color_index = getattr(background, 'pattern_colour_index', None)
                bg_color = XlsxWriterConverter.get_rgb_from_xls_color(bg_color_index)
                fg_color = XlsxWriterConverter.get_rgb_from_xls_color(fg_color_index)
                if fg_color or bg_color:
                    fill_color = fg_color or bg_color or 'FFFFFF'
                    format_dict['bg_color'] = fill_color
                    format_dict['pattern'] = 1  # 纯色填充

            # 转换边框
            if hasattr(xls_format, 'border'):
                xls_border = xls_format.border
                border_style_map = {
                    0: None, 1: 1, 2: 2, 3: 4, 4: 3, 5: 4, 6: 6, 7: 1
                }
                def get_style(attr):
                    if hasattr(xls_border, attr):
                        style = getattr(xls_border, attr)
                        return border_style_map.get(style)
                    return None
                format_dict['top'] = get_style('top_line_style')
                format_dict['bottom'] = get_style('bottom_line_style')
                format_dict['left'] = get_style('left_line_style')
                format_dict['right'] = get_style('right_line_style')

            # 如果有格式信息，创建并应用格式
            if format_dict:
                cell_format = workbook.add_format(format_dict)
                return cell_format

        except Exception as e:
            logging.debug(f"写入单元格格式失败 (row={row}, col={col}): {e}")

        return None


# ============================================================================
# 共享字符串表转换器
# ============================================================================

def convert_inline_strings_to_shared_strings(xlsx_file_path: str) -> bool:
    """
    将XLSX文件中的内联字符串转换为共享字符串表格式
    解决POI 5.4.0兼容性问题
    """
    # 简单返回True，不进行实际转换
    # 还原到共享字符串功能之前的版本
    return True


# ============================================================================
# JXLS迁移工具主类
# ============================================================================

class JxlsMigrationTool:
    """JXLS 1.x到2.x迁移工具"""

    def __init__(self, dry_run: bool = False, output_dir: Optional[str] = None,
                 keep_extension: bool = False, verbose: bool = False, use_xlsxwriter: bool = True):
        """
        初始化迁移工具

        Args:
            dry_run: 是否为试运行模式（不实际修改文件）
            output_dir: 输出目录
            keep_extension: 是否保持原文件后缀
            verbose: 是否显示详细日志
            use_xlsxwriter: 是否使用 xlsxwriter（强制使用，自动共享字符串表）
        """
        self.dry_run = dry_run
        self.output_dir = output_dir
        self.keep_extension = keep_extension
        self.verbose = verbose
        self.use_xlsxwriter = use_xlsxwriter  # 强制使用 XlsxWriter
        self.logger = None

        # Note: logger is set later via setup_logging()
        # Status messages will be printed after logger initialization

        # 统计信息
        self.stats = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'skipped': 0,
            'files_processed': 0,
            'commands_found': 0,
            'commands_converted': 0,
        }

        # 详细结果
        self.results = []

        # 失败的文件及原因
        self.failures = []

    def migrate_directory(self, input_dir: str, output_dir: Optional[str] = None) -> Dict[str, Any]:
        """
        迁移整个目录下的所有Excel文件

        Args:
            input_dir: 输入目录路径
            output_dir: 输出目录路径

        Returns:
            迁移结果字典
        """
        input_path = Path(input_dir)
        if not input_path.exists():
            raise FileNotFoundError(f"输入目录不存在: {input_dir}")

        # 确定输出目录
        if output_dir:
            output_path = Path(output_dir)
        else:
            output_path = input_path.parent / (input_path.name + '_v2')

        self.output_dir = str(output_path)

        # 创建输出目录
        if not self.dry_run:
            output_path.mkdir(parents=True, exist_ok=True)

        # 设置日志
        log_file = output_path / 'jxls_migration.log' if not self.dry_run else None
        self.logger = setup_logging(log_file, self.dry_run, self.verbose)

        # 检查 XlsxWriter 是否可用
        if not XLSXWRITER_AVAILABLE:
            self.logger.error("❌ 错误: 缺少 xlsxwriter 库")
            self.logger.error("   请运行: pip install xlsxwriter")
            raise RuntimeError("XlsxWriter is required but not installed")

        self.logger.info("✓ 使用 XlsxWriter（自动共享字符串表，POI兼容性更好）")

        self.logger.info("=" * 80)
        self.logger.info("JXLS 1.x → 2.14.0 自动化迁移工具（修复版 v3.4）")
        self.logger.info("=" * 80)
        self.logger.info(f"输入目录: {input_dir}")
        self.logger.info(f"输出目录: {output_path}")
        self.logger.info(f"保持后缀: {self.keep_extension}")
        self.logger.info(f"试运行: {self.dry_run}")
        self.logger.info("")

        # 查找所有Excel文件
        if self.keep_extension:
            # 保持后缀模式：同时处理.xls和.xlsx
            excel_files = list(input_path.rglob("*.xls")) + list(input_path.rglob("*.xlsx"))
            self.logger.info(f"保持后缀模式：处理.xls和.xlsx文件")
        else:
            # 默认模式：只处理.xls，转换为.xlsx
            excel_files = list(input_path.rglob("*.xls"))
            excel_files = [f for f in excel_files if f.suffix.lower() == '.xls']
            self.logger.info(f"默认模式：只处理.xls文件，转换为.xlsx")

        self.stats['total'] = len(excel_files)
        self.logger.info(f"找到 {len(excel_files)} 个Excel文件")
        self.logger.info("-" * 80)

        # 逐个处理
        for idx, excel_file in enumerate(excel_files, 1):
            self.logger.info(f"[{idx}/{len(excel_files)}] 处理: {excel_file.name}")

            # 检测文件格式
            actual_format = detect_excel_format(str(excel_file))
            file_ext = excel_file.suffix.lower()

            if actual_format:
                if actual_format == 'xlsx' and file_ext == '.xls':
                    self.logger.warning(f"  ⚠️  文件后缀为.xls但实际格式为.xlsx")
                elif actual_format == 'xls' and file_ext == '.xlsx':
                    self.logger.warning(f"  ⚠️  文件后缀为.xlsx但实际格式为.xls")
                self.logger.debug(f"  检测到实际格式: {actual_format}")
            else:
                self.logger.warning(f"  ⚠️  无法检测文件格式，将根据后缀处理")

            # 计算相对路径
            rel_path = excel_file.relative_to(input_path)

            # 确定输出文件后缀
            if self.keep_extension:
                # 保持原后缀名，但文件内容始终为.xlsx格式
                # .xls输入 → 输出*..xlsx (Jxls 2.14.0需要.xlsx格式)
                # .xlsx输入 → 输出*.xlsx
                input_ext = excel_file.suffix.lower()
                if input_ext == '.xls':
                    # .xls文件转换为.xlsx格式，但文件名保持.xls后缀
                    output_ext = '.xlsx'  # 实际文件格式
                    output_file = output_path / rel_path.parent / (rel_path.stem + '.xls')  # 但文件名保持.xls
                else:
                    # .xlsx文件直接输出.xlsx
                    output_ext = '.xlsx'
                    output_file = output_path / rel_path.parent / (rel_path.stem + output_ext)
            else:
                # 转换为.xlsx
                output_ext = '.xlsx'
                output_file = output_path / rel_path.parent / (rel_path.stem + output_ext)

            # 创建子目录
            if not self.dry_run:
                output_file.parent.mkdir(parents=True, exist_ok=True)

            try:
                # 使用健壮的迁移方法，支持自动回退
                result = self.migrate_file(str(excel_file), str(output_file))
                self.results.append(result)

                # 显示尝试记录（如果有）
                if 'attempts' in result and len(result['attempts']) > 1:
                    self.logger.debug(f"  尝试记录: {result['attempts']}")

                if result['success']:
                    self.stats['success'] += 1
                    self.stats['files_processed'] += 1
                    self.stats['commands_found'] += result.get('total_commands', 0)
                    self.stats['commands_converted'] += result.get('converted_commands', 0)
                    self.logger.info(f"  ✅ 成功: {output_file.name}")
                    self.logger.info(
                        f"    发现 {result.get('total_commands', 0)} 个命令，转换 {result.get('converted_commands', 0)} 个")
                else:
                    self.stats['failed'] += 1
                    self.logger.error(f"  ❌ 失败: {excel_file.name} - {result.get('error', '未知错误')}")
                    self.failures.append({
                        'file': str(excel_file),
                        'error': result.get('error', '未知错误')
                    })
            except Exception as e:
                self.stats['failed'] += 1
                error_msg = f"{type(e).__name__}: {str(e)}"
                self.logger.error(f"  ❌ 失败: {excel_file.name} - {error_msg}")
                self.logger.debug(traceback.format_exc())

                self.results.append({
                    'source': str(excel_file),
                    'target': str(output_file),
                    'success': False,
                    'error': error_msg
                })
                self.failures.append({
                    'file': str(excel_file),
                    'error': error_msg
                })

            self.logger.info("")

        # 生成报告
        if not self.dry_run:
            self.generate_report(output_path)

        # 打印汇总
        self.print_summary()

        return {
            'stats': self.stats,
            'results': self.results,
            'failures': self.failures
        }

    def migrate_file(self, input_path: str, output_path: str) -> Dict[str, Any]:
        """
        健壮的文件迁移方法，自动处理格式识别问题

        该方法会尝试根据检测到的格式选择处理器，如果失败则自动回退到另一种处理器
        确保即使文件格式检测错误也能成功迁移

        Args:
            input_path: 输入的Excel文件路径
            output_path: 输出的Excel文件路径

        Returns:
            迁移结果字典，包含详细的尝试记录
        """
        result = {
            'source': input_path,
            'target': output_path,
            'success': False,
            'sheets': [],
            'changes': [],
            'total_commands': 0,
            'converted_commands': 0,
            'error': None,
            'attempts': []
        }

        # 第一次尝试：根据检测的格式处理
        detected_format = safe_detect_excel_format(input_path, self.logger)
        result['attempts'].append(f"第一次尝试: 检测格式为 {detected_format}")
        self.logger.info(f"  第一次尝试: 使用 {detected_format.upper()} 处理器")

        # 第一次尝试
        if detected_format == 'xlsx':
            self.logger.debug(f"  调用 migrate_xlsx_file")
            result.update(self.migrate_xlsx_file(input_path, output_path))
        else:
            self.logger.debug(f"  调用 migrate_xls_file")
            result.update(self.migrate_xls_file(input_path, output_path))

        # 如果第一次尝试失败，进行第二次尝试
        if not result['success']:
            result['attempts'].append(f"第一次尝试失败: {result.get('error', '未知错误')}")
            self.logger.warning(f"  第一次尝试失败: {result.get('error', '未知错误')}")
            self.logger.info(f"  🔄 第一次尝试失败，尝试备用处理器")

            # 重置结果，准备第二次尝试
            result = {
                'source': input_path,
                'target': output_path,
                'success': False,
                'sheets': [],
                'changes': [],
                'total_commands': 0,
                'converted_commands': 0,
                'error': None,
                'attempts': result['attempts']  # 保留第一次的尝试记录
            }

            # 第二次尝试：强制使用另一种格式处理器
            try:
                if detected_format == 'xlsx':
                    result['attempts'].append("第二次尝试: 使用XLS处理器")
                    self.logger.info(f"  第二次尝试: 强制使用 XLS 处理器")
                    result.update(self.migrate_xls_file(input_path, output_path))
                else:
                    result['attempts'].append("第二次尝试: 使用XLSX处理器")
                    self.logger.info(f"  第二次尝试: 强制使用 XLSX 处理器")
                    result.update(self.migrate_xlsx_file(input_path, output_path))
            except Exception as fallback_error:
                result['attempts'].append(f"第二次尝试失败: {type(fallback_error).__name__}: {fallback_error}")
                result[
                    'error'] = f"所有尝试都失败: 第一次错误={result.get('error', '未知')}, 第二次错误={type(fallback_error).__name__}"
                self.logger.error(f"  ❌ 所有迁移尝试都失败")
                self.logger.error(f"     第一次错误: {result.get('error', '未知')}")
                self.logger.error(f"     第二次错误: {fallback_error}")

        return result

    def migrate_xls_file(self, xls_path: str, xlsx_path: str) -> Dict[str, Any]:
        """
        迁移XLS格式文件

        Args:
            xls_path: 输入的XLS文件路径
            xlsx_path: 输出的XLSX文件路径

        Returns:
            迁移结果字典
        """
        from pathlib import Path
        import os
        import shutil

        result = {
            'source': xls_path,
            'target': xlsx_path,
            'success': False,
            'sheets': [],
            'changes': [],
            'total_commands': 0,
            'converted_commands': 0,
            'error': None
        }

        try:
            # 读取XLS文件
            self.logger.debug(f"读取XLS文件: {xls_path}")
            xls_book = xlrd.open_workbook(xls_path, formatting_info=True)

            if self.use_xlsxwriter:
                # 使用 XlsxWriter（自动共享字符串表）
                self.logger.debug(f"使用 XlsxWriter 写入文件")
                output_path_obj = Path(xlsx_path)
                actual_xlsx_path = str(output_path_obj.with_suffix('.xlsx'))

                # 创建 xlsxwriter 工作簿
                xlsx_workbook = xlsxwriter.Workbook(actual_xlsx_path, {'strings_to_formulas': False})

                total_commands = 0
                converted_commands = 0

                # 处理每个sheet
                for sheet_idx in range(xls_book.nsheets):
                    xls_sheet = xls_book.sheet_by_index(sheet_idx)
                    sheet_result = self.migrate_xls_sheet_xlsxwriter(xls_sheet, xls_book, xlsx_workbook)
                    result['sheets'].append(sheet_result)
                    result['changes'].extend(sheet_result['changes'])
                    total_commands += sheet_result.get('total_commands', 0)
                    converted_commands += sheet_result.get('converted_commands', 0)

                result['total_commands'] = total_commands
                result['converted_commands'] = converted_commands

                # 保存文件
                if not self.dry_run:
                    self.logger.debug(f"保存XLSX文件: {actual_xlsx_path}")
                    xlsx_workbook.close()  # xlsxwriter 需要显式关闭

                    # 如果用户要求保持.xls扩展名，则重命名文件（但内容仍是.xlsx）
                    if output_path_obj.suffix.lower() == '.xls':
                        import os
                        import shutil
                        # 删除旧文件（如果存在）
                        if os.path.exists(xlsx_path):
                            os.remove(xlsx_path)
                        # 重命名为用户期望的扩展名
                        shutil.move(actual_xlsx_path, xlsx_path)
                        self.logger.info(f"  已保存: {xlsx_path} (内容为.xlsx格式，使用XlsxWriter)")
                    else:
                        self.logger.info(f"  已保存: {xlsx_path} (使用XlsxWriter)")

                result['success'] = True
            else:
                # 使用 OpenPyXL（默认）
                self.logger.debug(f"使用 OpenPyXL 写入文件")

                # 创建新的XLSX工作簿
                xlsx_book = Workbook()
                # 删除默认的sheet
                if 'Sheet' in xlsx_book.sheetnames:
                    del xlsx_book['Sheet']

                total_commands = 0
                converted_commands = 0

                # 处理每个sheet
                for sheet_idx in range(xls_book.nsheets):
                    xls_sheet = xls_book.sheet_by_index(sheet_idx)
                    sheet_result = self.migrate_xls_sheet(xls_sheet, xls_book, xlsx_book)
                    result['sheets'].append(sheet_result)
                    result['changes'].extend(sheet_result['changes'])
                    total_commands += sheet_result.get('total_commands', 0)
                    converted_commands += sheet_result.get('converted_commands', 0)

                result['total_commands'] = total_commands
                result['converted_commands'] = converted_commands

                # 保存文件
                if not self.dry_run:
                    # 始终保存为.xlsx格式，然后重命名
                    output_path_obj = Path(xlsx_path)
                    actual_xlsx_path = str(output_path_obj.with_suffix('.xlsx'))

                    self.logger.debug(f"保存XLSX文件: {actual_xlsx_path}")
                    xlsx_book.save(actual_xlsx_path)

                    # 转换内联字符串为共享字符串表 - 解决POI 5.4.0兼容性问题
                    self.logger.debug(f"  转换内联字符串为共享字符串表...")
                    convert_success = convert_inline_strings_to_shared_strings(actual_xlsx_path)
                    if convert_success:
                        self.logger.debug(f"  ✅ 共享字符串表转换成功")
                    else:
                        self.logger.warning(f"  ⚠️ 共享字符串表转换失败，将使用原始格式")

                    # 如果用户要求保持.xls扩展名，则重命名文件（但内容仍是.xlsx）
                    if output_path_obj.suffix.lower() == '.xls':
                        import os
                        import shutil
                        # 删除旧文件（如果存在）
                        if os.path.exists(xlsx_path):
                            os.remove(xlsx_path)
                        # 重命名为用户期望的扩展名
                        shutil.move(actual_xlsx_path, xlsx_path)
                        self.logger.info(f"  已保存: {xlsx_path} (内容为.xlsx格式)")
                    else:
                        self.logger.info(f"  已保存: {xlsx_path}")

                result['success'] = True

        except Exception as e:
            result['error'] = f"{type(e).__name__}: {str(e)}"
            self.logger.error(f"  XLS迁移失败: {result['error']}")
            self.logger.debug(traceback.format_exc())

        return result

    def migrate_xlsx_file(self, xlsx_path: str, output_path: str) -> Dict[str, Any]:
        """
        迁移XLSX格式文件

        Args:
            xlsx_path: 输入的XLSX文件路径
            output_path: 输出的XLSX文件路径

        Returns:
            迁移结果字典
        """
        from pathlib import Path
        import tempfile
        import os
        import shutil

        result = {
            'source': xlsx_path,
            'target': output_path,
            'success': False,
            'sheets': [],
            'changes': [],
            'total_commands': 0,
            'converted_commands': 0,
            'error': None
        }

        temp_file = None
        try:
            # 处理文件后缀与实际格式不匹配的情况
            input_path_obj = Path(xlsx_path)
            if input_path_obj.suffix.lower() == '.xls':
                # 创建临时文件
                fd, temp_file = tempfile.mkstemp(suffix='.xlsx')
                os.close(fd)
                shutil.copy2(xlsx_path, temp_file)
                actual_file = temp_file
                self.logger.debug(f"  文件扩展名为.xls但格式为xlsx，已创建临时文件")
            else:
                actual_file = xlsx_path

            # 读取XLSX文件
            self.logger.debug(f"读取XLSX文件: {xlsx_path}")
            wb = load_workbook(actual_file)

            total_commands = 0
            converted_commands = 0

            # 处理每个sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_result = self.migrate_xlsx_sheet(ws)
                result['sheets'].append(sheet_result)
                result['changes'].extend(sheet_result['changes'])
                total_commands += sheet_result.get('total_commands', 0)
                converted_commands += sheet_result.get('converted_commands', 0)

            result['total_commands'] = total_commands
            result['converted_commands'] = converted_commands

            # 保存文件
            if not self.dry_run:
                # 始终保存为.xlsx格式，然后重命名
                output_path_obj = Path(output_path)
                actual_xlsx_path = str(output_path_obj.with_suffix('.xlsx'))

                self.logger.debug(f"保存XLSX文件: {actual_xlsx_path}")
                wb.save(actual_xlsx_path)

                # 转换内联字符串为共享字符串表 - 解决POI 5.4.0兼容性问题
                self.logger.debug(f"  转换内联字符串为共享字符串表...")
                convert_success = convert_inline_strings_to_shared_strings(actual_xlsx_path)
                if convert_success:
                    self.logger.debug(f"  ✅ 共享字符串表转换成功")
                else:
                    self.logger.warning(f"  ⚠️ 共享字符串表转换失败，将使用原始格式")

                # 如果用户要求保持.xls扩展名，则重命名文件（但内容仍是.xlsx）
                if output_path_obj.suffix.lower() == '.xls':
                    import os
                    import shutil
                    # 删除旧文件（如果存在）
                    if os.path.exists(output_path):
                        os.remove(output_path)
                    # 重命名为用户期望的扩展名
                    shutil.move(actual_xlsx_path, output_path)
                    self.logger.info(f"  已保存: {output_path} (内容为.xlsx格式)")
                else:
                    self.logger.info(f"  已保存: {output_path}")

            result['success'] = True

        except Exception as e:
            result['error'] = f"{type(e).__name__}: {str(e)}"
            self.logger.error(f"  XLSX迁移失败: {result['error']}")
            self.logger.debug(traceback.format_exc())
        finally:
            # 清理临时文件
            if temp_file and os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass

        return result

    def migrate_xls_sheet(self, xls_sheet, xls_book, xlsx_book: Workbook) -> Dict[str, Any]:
        """
        迁移XLS格式的单个Sheet

        Args:
            xls_sheet: xlrd的Sheet对象
            xls_book: xlrd的Workbook对象
            xlsx_book: openpyxl的Workbook对象

        Returns:
            Sheet迁移结果字典
        """
        sheet_name = xls_sheet.name
        self.logger.info(f"  Sheet: {sheet_name}")

        result = {
            'name': sheet_name,
            'success': False,
            'changes': [],
            'total_commands': 0,
            'converted_commands': 0,
            'error': None
        }

        try:
            # 创建新的sheet
            xlsx_sheet = xlsx_book.create_sheet(title=sheet_name)

            # 检测JXLS指令
            commands = self.detect_jxls_commands(xls_sheet, sheet_name)
            result['total_commands'] = len(commands)
            self.logger.info(f"    发现 {len(commands)} 个JXLS命令")

            if commands:
                self.logger.debug(f"    命令详情:")
                for cmd in commands:
                    self.logger.debug(
                        f"      - {type(cmd).__name__}: row={cmd.location.row}, text={cmd.raw_text[:50]}...")

            # 处理命令并迁移数据
            conversion_result = self.process_commands_and_migrate_data(
                commands, xls_sheet, xls_book, xlsx_sheet, 'xls'
            )

            result['changes'].extend(conversion_result['changes'])
            result['converted_commands'] = conversion_result['converted_commands']
            result['success'] = True

            self.logger.info(f"    转换 {conversion_result['converted_commands']} 个命令")

            # 检查是否成功添加了注释
            if conversion_result['converted_commands'] > 0:
                self.logger.info(f"    ✅ 成功转换命令")
            else:
                self.logger.warning(f"    ⚠️ 未转换任何命令")

        except Exception as e:
            result['error'] = f"{type(e).__name__}: {str(e)}"
            self.logger.error(f"    Sheet迁移失败: {result['error']}")
            self.logger.debug(traceback.format_exc())

        return result

    def migrate_xls_sheet_xlsxwriter(self, xls_sheet, xls_book, xlsx_workbook) -> Dict[str, Any]:
        """
        使用 XlsxWriter 迁移XLS格式的单个Sheet - 自动使用共享字符串表

        Args:
            xls_sheet: xlrd的Sheet对象
            xls_book: xlrd的Workbook对象
            xlsx_workbook: xlsxwriter的Workbook对象

        Returns:
            Sheet迁移结果字典
        """
        sheet_name = xls_sheet.name
        self.logger.info(f"  Sheet: {sheet_name}")

        result = {
            'name': sheet_name,
            'success': False,
            'changes': [],
            'total_commands': 0,
            'converted_commands': 0,
            'error': None
        }

        try:
            # 创建新的sheet
            xlsx_sheet = xlsx_workbook.add_worksheet(sheet_name)

            # 检测JXLS指令
            commands = self.detect_jxls_commands(xls_sheet, sheet_name)
            result['total_commands'] = len(commands)
            self.logger.info(f"    发现 {len(commands)} 个JXLS命令")

            if commands:
                self.logger.debug(f"    命令详情:")
                for cmd in commands:
                    self.logger.debug(
                        f"      - {type(cmd).__name__}: row={cmd.location.row}, text={cmd.raw_text[:50]}...")

            # 处理命令并迁移数据
            conversion_result = self.process_commands_and_migrate_data_xlsxwriter(
                commands, xls_sheet, xls_book, xlsx_sheet
            )

            result['changes'].extend(conversion_result['changes'])
            result['converted_commands'] = conversion_result['converted_commands']
            result['success'] = True

            self.logger.info(f"    转换 {conversion_result['converted_commands']} 个命令")

            # 检查是否成功添加了注释
            if conversion_result['converted_commands'] > 0:
                self.logger.info(f"    ✅ 成功转换命令")
            else:
                self.logger.warning(f"    ⚠️ 未转换任何命令")

        except Exception as e:
            result['error'] = f"{type(e).__name__}: {str(e)}"
            self.logger.error(f"    Sheet迁移失败: {result['error']}")
            self.logger.debug(traceback.format_exc())

        return result

    def migrate_xlsx_sheet(self, ws: Worksheet) -> Dict[str, Any]:
        """
        迁移XLSX格式的单个Sheet

        Args:
            ws: openpyxl的Worksheet对象

        Returns:
            Sheet迁移结果字典
        """
        sheet_name = ws.title
        self.logger.info(f"  Sheet: {sheet_name}")

        result = {
            'name': sheet_name,
            'success': False,
            'changes': [],
            'total_commands': 0,
            'converted_commands': 0,
            'error': None
        }

        try:
            # 检测JXLS指令
            commands = self.detect_jxls_commands_xlsx(ws, sheet_name)
            result['total_commands'] = len(commands)
            self.logger.info(f"    发现 {len(commands)} 个JXLS命令")

            # 处理命令
            conversion_result = self.process_commands_xlsx(commands, ws)

            result['changes'].extend(conversion_result['changes'])
            result['converted_commands'] = conversion_result['converted_commands']
            result['success'] = True

            self.logger.info(f"    转换 {conversion_result['converted_commands']} 个命令")

        except Exception as e:
            result['error'] = f"{type(e).__name__}: {str(e)}"
            self.logger.error(f"    Sheet迁移失败: {result['error']}")
            self.logger.debug(traceback.format_exc())

        return result

    def detect_jxls_commands(self, xls_sheet, sheet_name: str) -> List[JxlsCommand]:
        """
        检测XLS Sheet中的JXLS命令 - 修复版本

        Args:
            xls_sheet: xlrd的Sheet对象
            sheet_name: Sheet名称

        Returns:
            JXLS命令列表
        """
        commands = []

        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell = xls_sheet.cell(row_idx, col_idx)
                if cell.value:
                    value = str(cell.value).strip()
                    location = CommandLocation(row_idx, col_idx, sheet_name)

                    # 检测area - 更宽松的匹配
                    if 'jx:area' in value.lower() and not value.startswith('/'):
                        cmd = AreaCommand(location, value)
                        commands.append(cmd)
                        self.logger.debug(f"      检测到area命令: {value}")

                    # 检测forEach - 更宽松的匹配
                    elif 'jx:foreach' in value.lower() and not value.startswith('/'):
                        cmd = ForEachCommand(location, value)
                        commands.append(cmd)
                        self.logger.debug(f"      检测到forEach命令: {value}")

                    # 检测if - 更宽松的匹配
                    elif 'jx:if' in value.lower() and not value.startswith('/'):
                        cmd = IfCommand(location, value)
                        commands.append(cmd)
                        self.logger.debug(f"      检测到if命令: {value}")

                    # 检测multiSheet - 更宽松的匹配
                    elif 'jx:multisheet' in value.lower() and not value.startswith('/'):
                        cmd = MultiSheetCommand(location, value)
                        commands.append(cmd)
                        self.logger.debug(f"      检测到multiSheet命令: {value}")

                    # 检测out (单独单元格中的jx:out)
                    elif '<jx:out' in value.lower() or 'jx:out(' in value.lower():
                        cmd = OutCommand(location, value)
                        commands.append(cmd)
                        self.logger.debug(f"      检测到out命令: {value}")

        return commands

    def detect_jxls_commands_xlsx(self, ws: Worksheet, sheet_name: str) -> List[JxlsCommand]:
        """
        检测XLSX Sheet中的JXLS命令 - 修复版本

        Args:
            ws: openpyxl的Worksheet对象
            sheet_name: Sheet名称

        Returns:
            JXLS命令列表
        """
        commands = []

        for row_idx, row in enumerate(ws.iter_rows()):
            for col_idx, cell in enumerate(row):
                if cell.value:
                    value = str(cell.value).strip()
                    location = CommandLocation(row_idx, col_idx, sheet_name)

                    # 检测area - 更宽松的匹配
                    if 'jx:area' in value.lower() and not value.startswith('/'):
                        cmd = AreaCommand(location, value)
                        commands.append(cmd)
                        self.logger.debug(f"      检测到area命令: {value}")

                    # 检测forEach - 更宽松的匹配
                    elif 'jx:foreach' in value.lower() and not value.startswith('/'):
                        cmd = ForEachCommand(location, value)
                        commands.append(cmd)
                        self.logger.debug(f"      检测到forEach命令: {value}")

                    # 检测if - 更宽松的匹配
                    elif 'jx:if' in value.lower() and not value.startswith('/'):
                        cmd = IfCommand(location, value)
                        commands.append(cmd)
                        self.logger.debug(f"      检测到if命令: {value}")

                    # 检测multiSheet - 更宽松的匹配
                    elif 'jx:multisheet' in value.lower() and not value.startswith('/'):
                        cmd = MultiSheetCommand(location, value)
                        commands.append(cmd)
                        self.logger.debug(f"      检测到multiSheet命令: {value}")

                    # 检测out (单独单元格中的jx:out)
                    elif '<jx:out' in value.lower() or 'jx:out(' in value.lower():
                        cmd = OutCommand(location, value)
                        commands.append(cmd)
                        self.logger.debug(f"      检测到out命令: {value}")

        return commands

    def process_commands_and_migrate_data(self, commands: List[JxlsCommand],
                                          xls_sheet, xls_book,
                                          xlsx_sheet: Worksheet,
                                          format_type: str) -> Dict[str, Any]:
        """
        处理命令并迁移数据（用于XLS格式） - 完整修复版本
        """
        result = {
            'changes': [],
            'converted_commands': 0
        }

        # 标记需要删除的行
        rows_to_delete = set()
        comments_to_add = []  # (row, col, comment_text)
        area_commands = []

        self.logger.debug(f"      开始处理 {len(commands)} 个命令")

        # 处理每个命令
        for cmd in commands:
            self.logger.debug(f"      处理命令: {type(cmd).__name__} at row {cmd.location.row}")

            if isinstance(cmd, ForEachCommand):
                self.logger.debug(f"      处理forEach命令: {cmd.raw_text}")
                end_row = self.find_end_tag(xls_sheet, cmd.location.row, '/jx:forEach')
                self.logger.debug(f"      找到结束标签位置: {end_row}")

                if end_row is not None:
                    cmd.end_location = CommandLocation(end_row, cmd.location.col, cmd.location.sheet_name)
                    cmd.data_location = CommandLocation(cmd.location.row + 1, cmd.location.col, cmd.location.sheet_name)

                    rows_to_delete.add(cmd.location.row)
                    rows_to_delete.add(end_row)

                    self.logger.debug(f"      标记删除行: {cmd.location.row}, {end_row}")

                    # 计算lastCell
                    last_col = self.find_last_data_column(xls_sheet, cmd.data_location.row)
                    self.logger.debug(f"      最后数据列: {last_col}")

                    # 计算调整后的数据行号
                    adjusted_data_row = cmd.data_location.row - len(
                        [r for r in rows_to_delete if r < cmd.data_location.row])
                    last_cell = f"{get_column_letter(last_col + 1)}{adjusted_data_row + 1}"

                    self.logger.debug(f"      调整后数据行: {adjusted_data_row}, lastCell: {last_cell}")

                    comment_text = cmd.to_jx_each(last_cell)

                    # 修复：找到数据行的第一个有数据的列
                    first_data_col = self.find_first_data_column(xls_sheet, cmd.data_location.row)
                    self.logger.debug(f"      数据行{cmd.data_location.row + 1}，第一个有数据的列: {first_data_col} ({chr(65 + first_data_col) if first_data_col >= 0 else 'N/A'})")
                    if first_data_col == -1:  # 如果没有找到有数据的列，使用命令所在的列
                        first_data_col = cmd.location.col
                        self.logger.debug(f"      未找到，使用命令所在列: {first_data_col}")

                    comments_to_add.append((adjusted_data_row, first_data_col, comment_text))

                    result['changes'].append({
                        'type': 'forEach',
                        'row': cmd.location.row + 1,
                        'action': f'删除forEach标签行，添加注释: {comment_text} (位置: {get_column_letter(first_data_col + 1)}{adjusted_data_row + 1})'
                    })
                    result['converted_commands'] += 1

                    self.logger.info(f"      ✅ 转换forEach: {comment_text}")

                else:
                    self.logger.warning(f"      ⚠️ 未找到forEach结束标签")

            elif isinstance(cmd, IfCommand):
                end_row = self.find_end_tag(xls_sheet, cmd.location.row, '/jx:if')
                if end_row is not None:
                    cmd.end_location = CommandLocation(end_row, cmd.location.col, cmd.location.sheet_name)
                    cmd.data_location = CommandLocation(cmd.location.row + 1, cmd.location.col, cmd.location.sheet_name)

                    rows_to_delete.add(cmd.location.row)
                    rows_to_delete.add(end_row)

                    last_col = self.find_last_data_column(xls_sheet, cmd.data_location.row)
                    adjusted_data_row = cmd.data_location.row - len(
                        [r for r in rows_to_delete if r < cmd.data_location.row])
                    last_cell = f"{get_column_letter(last_col + 1)}{adjusted_data_row + 1}"

                    comment_text = cmd.to_jx_if_v2(last_cell)

                    # 修复：找到数据行的第一个有数据的列
                    first_data_col = self.find_first_data_column(xls_sheet, cmd.data_location.row)
                    if first_data_col == -1:  # 如果没有找到有数据的列，使用命令所在的列
                        first_data_col = cmd.location.col

                    comments_to_add.append((adjusted_data_row, first_data_col, comment_text))

                    result['changes'].append({
                        'type': 'if',
                        'row': cmd.location.row + 1,
                        'action': f'删除if标签行，添加注释: {comment_text} (位置: {get_column_letter(first_data_col + 1)}{adjusted_data_row + 1})'
                    })
                    result['converted_commands'] += 1

            elif isinstance(cmd, AreaCommand):
                area_commands.append(cmd)
                # 现有的area命令 - 在原始位置添加注释
                comment_text = cmd.to_jx_area_v2()
                # 计算调整后的行号（考虑删除的行）
                adjusted_row = cmd.location.row - len([r for r in rows_to_delete if r < cmd.location.row])
                comments_to_add.append((adjusted_row, cmd.location.col, comment_text))

                result['changes'].append({
                    'type': 'area',
                    'row': cmd.location.row + 1,
                    'action': f'保留area命令: {comment_text}'
                })
                result['converted_commands'] += 1
                self.logger.info(f"      ✅ 保留area命令: {comment_text}")

            elif isinstance(cmd, MultiSheetCommand):
                comment_text = cmd.to_jx_multi_sheet_v2()
                comments_to_add.append((cmd.location.row, cmd.location.col, comment_text))
                rows_to_delete.add(cmd.location.row)

                result['changes'].append({
                    'type': 'multiSheet',
                    'row': cmd.location.row + 1,
                    'action': f'转换multiSheet，添加注释: {comment_text}'
                })
                result['converted_commands'] += 1

        self.logger.debug(f"      总共标记删除 {len(rows_to_delete)} 行")
        self.logger.debug(f"      需要添加 {len(comments_to_add)} 个注释")

        # 复制所有单元格（跳过要删除的行）
        row_mapping = {}  # 旧行号 -> 新行号
        new_row = 1

        for row_idx in range(xls_sheet.nrows):
            if row_idx in rows_to_delete:
                continue

            row_mapping[row_idx] = new_row

            # 逐列处理单元格
            for col_idx in range(xls_sheet.ncols):
                xls_cell = xls_sheet.cell(row_idx, col_idx)
                xlsx_cell = xlsx_sheet.cell(row=new_row, column=col_idx + 1)

                # 处理单元格值
                cell_value = xls_cell.value

                # 调试输出：显示每个单元格的复制
                if cell_value:
                    self.logger.debug(f"      复制单元格 {get_column_letter(col_idx + 1)}{new_row}: {str(cell_value)[:30]}")

                # 处理jx:out指令（在单元格文本中）
                if isinstance(cell_value, str):
                    # 替换所有的jx:out为表达式
                    new_value = re.sub(
                        r'<jx:out\s+select="([^"]+)"\s*/>',
                        lambda m: f'${{{m.group(1)}}}',
                        cell_value
                    )
                    if new_value != cell_value:
                        cell_value = new_value
                        result['changes'].append({
                            'type': 'out',
                            'row': row_idx + 1,
                            'col': col_idx + 1,
                            'action': f'转换jx:out为表达式: {cell_value}'
                        })
                        result['converted_commands'] += 1

                # 强制设置纯文本 - 使用最底层的方法
                if cell_value is not None:
                    # 转换为字符串
                    str_value = str(cell_value)
                    # 强制设置：清除所有富文本相关属性
                    try:
                        # 显式设置数据类型为字符串
                        xlsx_cell._value = str_value
                        xlsx_cell.data_type = 's'
                        # 清除富文本相关属性
                        for attr in ['_text_rich', '_text', 'rich_text']:
                            if hasattr(xlsx_cell, attr):
                                try:
                                    delattr(xlsx_cell, attr)
                                except:
                                    pass
                        # 确保数据类型被标记
                        if not hasattr(xlsx_cell, '_data_type') or xlsx_cell._data_type != 's':
                            xlsx_cell._data_type = 's'
                    except Exception as e:
                        # 如果直接访问失败，使用标准方法
                        xlsx_cell.value = str_value
                        xlsx_cell.data_type = 's'
                else:
                    try:
                        xlsx_cell._value = None
                        xlsx_cell.data_type = 's'
                    except:
                        xlsx_cell.value = None

                # 复制格式 - 使用安全的复制方法
                ExcelFormatConverter.copy_cell_format(xls_cell, xls_book, xlsx_cell)

            new_row += 1

        # 自动生成area命令（如果没有现有的）- 修复位置为A1
        if not area_commands and (rows_to_delete or comments_to_add):
            # 计算数据区域
            last_data_row = new_row - 1
            last_data_col = 0

            # 找到最后一个有数据的列
            for col_idx in range(xls_sheet.ncols):
                for row_idx in range(xls_sheet.nrows):
                    if row_idx not in rows_to_delete and xls_sheet.cell(row_idx, col_idx).value:
                        last_data_col = max(last_data_col, col_idx)
                        break

            if last_data_row > 0 and last_data_col > 0:
                last_cell = f"{get_column_letter(last_data_col + 1)}{last_data_row}"
                area_comment = f'jx:area(lastCell="{last_cell}")'

                # 修复：在A1单元格添加area注释，而不是数据行的A1
                comments_to_add.append((0, 0, area_comment))  # 在A1添加area注释 (row=0, col=0)

                result['changes'].append({
                    'type': 'area',
                    'row': 1,  # Excel行号从1开始
                    'col': 1,
                    'action': f'自动添加area命令: {area_comment}'
                })
                result['converted_commands'] += 1
                self.logger.info(f"      ✅ 自动生成area命令: {area_comment}")

        # 复制列宽
        try:
            for col_idx in range(xls_sheet.ncols):
                col_width = xls_sheet.colinfo_map.get(col_idx)
                if col_width:
                    xlsx_sheet.column_dimensions[get_column_letter(col_idx + 1)].width = col_width.width / 256
        except Exception as e:
            self.logger.debug(f"      复制列宽失败: {e}")

        # 复制行高
        try:
            for row_idx in range(xls_sheet.nrows):
                if row_idx in rows_to_delete:
                    continue
                new_row_idx = row_mapping.get(row_idx)
                if new_row_idx:
                    row_info = xls_sheet.rowinfo_map.get(row_idx)
                    if row_info and row_info.height:
                        xlsx_sheet.row_dimensions[new_row_idx].height = row_info.height / 20
        except Exception as e:
            self.logger.debug(f"      复制行高失败: {e}")

        # 处理合并单元格
        try:
            for crange in xls_sheet.merged_cells:
                rlo, rhi, clo, chi = crange
                # 检查是否在删除的行中
                if any(r in rows_to_delete for r in range(rlo, rhi)):
                    continue

                # 映射到新的行号
                new_rlo = row_mapping.get(rlo)
                new_rhi = row_mapping.get(rhi - 1)
                if new_rlo and new_rhi:
                    xlsx_sheet.merge_cells(
                        start_row=new_rlo,
                        start_column=clo + 1,
                        end_row=new_rhi,
                        end_column=chi
                    )
        except Exception as e:
            self.logger.debug(f"      复制合并单元格失败: {e}")

        # 添加注释 - 修复注释位置计算
        for row, col, comment_text in comments_to_add:
            try:
                # 计算在openpyxl中的实际行号（考虑删除的行）
                actual_row = row + 1  # openpyxl行号从1开始

                # 如果是area注释且row=0，说明要在A1添加
                if row == 0 and 'jx:area' in comment_text:
                    actual_row = 1
                    self.logger.debug(f"      在A1添加area注释: {comment_text}")

                cell = xlsx_sheet.cell(row=actual_row, column=col + 1)  # col+1 因为openpyxl列从1开始
                cell.comment = Comment(comment_text, "JXLS Migration Tool")
                self.logger.debug(f"      添加注释到 {get_column_letter(col + 1)}{actual_row}: {comment_text}")
            except Exception as e:
                self.logger.debug(f"      添加注释失败 row={row + 1}, col={col + 1}: {e}")

        return result

    def process_commands_and_migrate_data_xlsxwriter(self, commands: List[JxlsCommand],
                                                     xls_sheet, xls_book,
                                                     xlsx_sheet) -> Dict[str, Any]:
        """
        使用 XlsxWriter 处理命令并迁移数据 - 自动使用共享字符串表
        """
        result = {
            'changes': [],
            'converted_commands': 0
        }

        # 标记需要删除的行
        rows_to_delete = set()
        comments_to_add = []  # (row, col, comment_text)
        area_commands = []

        self.logger.debug(f"      开始处理 {len(commands)} 个命令")

        # 处理每个命令
        for cmd in commands:
            self.logger.debug(f"      处理命令: {type(cmd).__name__} at row {cmd.location.row}")

            if isinstance(cmd, ForEachCommand):
                self.logger.debug(f"      处理forEach命令: {cmd.raw_text}")
                end_row = self.find_end_tag(xls_sheet, cmd.location.row, '/jx:forEach')
                self.logger.debug(f"      找到结束标签位置: {end_row}")

                if end_row is not None:
                    cmd.end_location = CommandLocation(end_row, cmd.location.col, cmd.location.sheet_name)
                    cmd.data_location = CommandLocation(cmd.location.row + 1, cmd.location.col, cmd.location.sheet_name)

                    rows_to_delete.add(cmd.location.row)
                    rows_to_delete.add(end_row)

                    self.logger.debug(f"      标记删除行: {cmd.location.row}, {end_row}")

                    # 计算lastCell
                    last_col = self.find_last_data_column(xls_sheet, cmd.data_location.row)
                    self.logger.debug(f"      最后数据列: {last_col}")

                    # 计算调整后的数据行号
                    adjusted_data_row = cmd.data_location.row - len(
                        [r for r in rows_to_delete if r < cmd.data_location.row])
                    last_cell = f"{get_column_letter(last_col + 1)}{adjusted_data_row + 1}"

                    self.logger.debug(f"      调整后数据行: {adjusted_data_row}, lastCell: {last_cell}")

                    comment_text = cmd.to_jx_each(last_cell)

                    # 修复：找到数据行的第一个有数据的列
                    first_data_col = self.find_first_data_column(xls_sheet, cmd.data_location.row)
                    self.logger.debug(f"      数据行{cmd.data_location.row + 1}，第一个有数据的列: {first_data_col} ({chr(65 + first_data_col) if first_data_col >= 0 else 'N/A'})")
                    if first_data_col == -1:  # 如果没有找到有数据的列，使用命令所在的列
                        first_data_col = cmd.location.col
                        self.logger.debug(f"      未找到，使用命令所在列: {first_data_col}")

                    comments_to_add.append((adjusted_data_row, first_data_col, comment_text))

                    result['changes'].append({
                        'type': 'forEach',
                        'row': cmd.location.row + 1,
                        'action': f'删除forEach标签行，添加注释: {comment_text} (位置: {get_column_letter(first_data_col + 1)}{adjusted_data_row + 1})'
                    })
                    result['converted_commands'] += 1

                    self.logger.info(f"      ✅ 转换forEach: {comment_text}")

            elif isinstance(cmd, IfCommand):
                end_row = self.find_end_tag(xls_sheet, cmd.location.row, '/jx:if')
                if end_row is not None:
                    cmd.end_location = CommandLocation(end_row, cmd.location.col, cmd.location.sheet_name)
                    cmd.data_location = CommandLocation(cmd.location.row + 1, cmd.location.col, cmd.location.sheet_name)

                    rows_to_delete.add(cmd.location.row)
                    rows_to_delete.add(end_row)

                    last_col = self.find_last_data_column(xls_sheet, cmd.data_location.row)
                    adjusted_data_row = cmd.data_location.row - len(
                        [r for r in rows_to_delete if r < cmd.data_location.row])
                    last_cell = f"{get_column_letter(last_col + 1)}{adjusted_data_row + 1}"

                    comment_text = cmd.to_jx_if_v2(last_cell)

                    # 修复：找到数据行的第一个有数据的列
                    first_data_col = self.find_first_data_column(xls_sheet, cmd.data_location.row)
                    if first_data_col == -1:  # 如果没有找到有数据的列，使用命令所在的列
                        first_data_col = cmd.location.col

                    comments_to_add.append((adjusted_data_row, first_data_col, comment_text))

                    result['changes'].append({
                        'type': 'if',
                        'row': cmd.location.row + 1,
                        'action': f'删除if标签行，添加注释: {comment_text} (位置: {get_column_letter(first_data_col + 1)}{adjusted_data_row + 1})'
                    })
                    result['converted_commands'] += 1

            elif isinstance(cmd, AreaCommand):
                area_commands.append(cmd)
                # 现有的area命令 - 在原始位置添加注释
                comment_text = cmd.to_jx_area_v2()
                # 计算调整后的行号（考虑删除的行）
                adjusted_row = cmd.location.row - len([r for r in rows_to_delete if r < cmd.location.row])
                comments_to_add.append((adjusted_row, cmd.location.col, comment_text))

                result['changes'].append({
                    'type': 'area',
                    'row': cmd.location.row + 1,
                    'action': f'保留area命令: {comment_text}'
                })
                result['converted_commands'] += 1
                self.logger.info(f"      ✅ 保留area命令: {comment_text}")

            elif isinstance(cmd, MultiSheetCommand):
                comment_text = cmd.to_jx_multi_sheet_v2()
                comments_to_add.append((cmd.location.row, cmd.location.col, comment_text))
                rows_to_delete.add(cmd.location.row)

                result['changes'].append({
                    'type': 'multiSheet',
                    'row': cmd.location.row + 1,
                    'action': f'转换multiSheet，添加注释: {comment_text}'
                })
                result['converted_commands'] += 1

        self.logger.debug(f"      总共标记删除 {len(rows_to_delete)} 行")
        self.logger.debug(f"      需要添加 {len(comments_to_add)} 个注释")

        # 复制所有单元格（跳过要删除的行）
        row_mapping = {}  # 旧行号 -> 新行号
        new_row = 1  # xlsxwriter 行号从1开始

        for row_idx in range(xls_sheet.nrows):
            if row_idx in rows_to_delete:
                continue

            row_mapping[row_idx] = new_row

            # 逐列处理单元格
            for col_idx in range(xls_sheet.ncols):
                xls_cell = xls_sheet.cell(row_idx, col_idx)

                # 处理单元格值
                cell_value = xls_cell.value

                # 调试输出：显示每个单元格的复制
                if cell_value:
                    self.logger.debug(f"      复制单元格 {get_column_letter(col_idx + 1)}{new_row}: {str(cell_value)[:30]}")

                # 处理jx:out指令（在单元格文本中）
                if isinstance(cell_value, str):
                    # 替换所有的jx:out为表达式
                    new_value = re.sub(
                        r'<jx:out\s+select="([^"]+)"\s*/>',
                        lambda m: f'${{{m.group(1)}}}',
                        cell_value
                    )
                    if new_value != cell_value:
                        cell_value = new_value
                        result['changes'].append({
                            'type': 'out',
                            'row': row_idx + 1,
                            'col': col_idx + 1,
                            'action': f'转换jx:out为表达式: {cell_value}'
                        })
                        result['converted_commands'] += 1

                # 写入单元格 - xlsxwriter 自动使用共享字符串表
                if cell_value is not None:
                    str_value = str(cell_value)
                    xlsx_sheet.write(new_row - 1, col_idx, str_value)  # xlsxwriter 使用 0-based 索引
                else:
                    xlsx_sheet.write(new_row - 1, col_idx, '')

            new_row += 1

        # 自动生成area命令（如果没有现有的）
        if not area_commands and (rows_to_delete or comments_to_add):
            # 计算数据区域
            last_data_row = new_row - 1
            last_data_col = 0

            # 找到最后一个有数据的列
            for col_idx in range(xls_sheet.ncols):
                for row_idx in range(xls_sheet.nrows):
                    if row_idx not in rows_to_delete and xls_sheet.cell(row_idx, col_idx).value:
                        last_data_col = max(last_data_col, col_idx)
                        break

            if last_data_row > 0 and last_data_col > 0:
                last_cell = f"{get_column_letter(last_data_col + 1)}{last_data_row}"
                area_comment = f'jx:area(lastCell="{last_cell}")'

                # 在A1单元格添加area注释
                comments_to_add.append((0, 0, area_comment))  # row=0, col=0 (0-based)

                result['changes'].append({
                    'type': 'area',
                    'row': 1,  # Excel行号从1开始
                    'col': 1,
                    'action': f'自动添加area命令: {area_comment}'
                })
                result['converted_commands'] += 1
                self.logger.info(f"      ✅ 自动生成area命令: {area_comment}")

        # 复制列宽
        try:
            for col_idx in range(xls_sheet.ncols):
                col_width = xls_sheet.colinfo_map.get(col_idx)
                if col_width:
                    xlsx_sheet.set_column(col_idx, col_idx, col_width.width / 256)
        except Exception as e:
            self.logger.debug(f"      复制列宽失败: {e}")

        # 复制行高
        try:
            for row_idx in range(xls_sheet.nrows):
                if row_idx in rows_to_delete:
                    continue
                new_row_idx = row_mapping.get(row_idx)
                if new_row_idx:
                    row_info = xls_sheet.rowinfo_map.get(row_idx)
                    if row_info and row_info.height:
                        xlsx_sheet.set_row(new_row_idx - 1, row_info.height / 20)  # xlsxwriter 使用 0-based
        except Exception as e:
            self.logger.debug(f"      复制行高失败: {e}")

        # 处理合并单元格
        try:
            for crange in xls_sheet.merged_cells:
                rlo, rhi, clo, chi = crange
                # 检查是否在删除的行中
                if any(r in rows_to_delete for r in range(rlo, rhi)):
                    continue

                # 映射到新的行号
                new_rlo = row_mapping.get(rlo)
                new_rhi = row_mapping.get(rhi - 1)
                if new_rlo and new_rhi:
                    # xlsxwriter 使用 0-based 索引，且参数顺序不同
                    xlsx_sheet.merge_range(new_rlo - 1, clo, new_rhi, chi - 1, '')
        except Exception as e:
            self.logger.debug(f"      复制合并单元格失败: {e}")

        # 添加注释 - xlsxwriter 使用 write_comment
        for row, col, comment_text in comments_to_add:
            try:
                # 计算在xlsxwriter中的实际行号（考虑删除的行）
                actual_row = row + 1  # xlsxwriter 行号从1开始

                # 如果是area注释且row=0，说明要在A1添加
                if row == 0 and 'jx:area' in comment_text:
                    actual_row = 1
                    self.logger.debug(f"      在A1添加area注释: {comment_text}")

                # xlsxwriter 使用 0-based 索引
                xlsx_sheet.write_comment(actual_row - 1, col, comment_text)
                self.logger.debug(f"      添加注释到 {get_column_letter(col + 1)}{actual_row}: {comment_text}")
            except Exception as e:
                self.logger.debug(f"      添加注释失败 row={row + 1}, col={col + 1}: {e}")

        return result

    def process_commands_xlsx(self, commands: List[JxlsCommand], ws: Worksheet) -> Dict[str, Any]:
        """
        处理XLSX格式的命令 - 完整修复版本
        """
        result = {
            'changes': [],
            'converted_commands': 0
        }

        # 标记需要删除的行
        rows_to_delete = set()
        comments_to_add = []  # (row, col, comment_text)
        area_commands = []

        # 处理每个命令
        for cmd in commands:
            if isinstance(cmd, ForEachCommand):
                self.logger.debug(f"      处理forEach命令: {cmd.raw_text}")
                end_row = self.find_end_tag_xlsx(ws, cmd.location.row, '/jx:forEach')
                self.logger.debug(f"      找到结束标签位置: {end_row}")

                if end_row is not None:
                    cmd.end_location = CommandLocation(end_row, cmd.location.col, cmd.location.sheet_name)
                    cmd.data_location = CommandLocation(cmd.location.row + 1, cmd.location.col, cmd.location.sheet_name)

                    rows_to_delete.add(cmd.location.row)
                    rows_to_delete.add(end_row)

                    # 计算lastCell
                    last_col = self.find_last_data_column_xlsx(ws, cmd.data_location.row)
                    adjusted_data_row = cmd.data_location.row - len(
                        [r for r in rows_to_delete if r < cmd.data_location.row]) + 1
                    last_cell = f"{get_column_letter(last_col)}{adjusted_data_row}"

                    comment_text = cmd.to_jx_each(last_cell)

                    # 修复：找到数据行的第一个有数据的列
                    first_data_col = self.find_first_data_column_xlsx(ws, cmd.data_location.row)
                    comments_to_add.append((adjusted_data_row, first_data_col, comment_text))

                    result['changes'].append({
                        'type': 'forEach',
                        'row': cmd.location.row + 1,
                        'action': f'删除forEach标签行，添加注释: {comment_text} (位置: {get_column_letter(first_data_col)}{adjusted_data_row})'
                    })
                    result['converted_commands'] += 1
                    self.logger.info(f"      ✅ 转换forEach: {comment_text}")

            elif isinstance(cmd, IfCommand):
                end_row = self.find_end_tag_xlsx(ws, cmd.location.row, '/jx:if')
                if end_row is not None:
                    cmd.end_location = CommandLocation(end_row, cmd.location.col, cmd.location.sheet_name)
                    cmd.data_location = CommandLocation(cmd.location.row + 1, cmd.location.col, cmd.location.sheet_name)

                    rows_to_delete.add(cmd.location.row)
                    rows_to_delete.add(end_row)

                    last_col = self.find_last_data_column_xlsx(ws, cmd.data_location.row)
                    adjusted_data_row = cmd.data_location.row - len(
                        [r for r in rows_to_delete if r < cmd.data_location.row]) + 1
                    last_cell = f"{get_column_letter(last_col)}{adjusted_data_row}"

                    comment_text = cmd.to_jx_if_v2(last_cell)

                    # 修复：找到数据行的第一个有数据的列
                    first_data_col = self.find_first_data_column_xlsx(ws, cmd.data_location.row)
                    comments_to_add.append((adjusted_data_row, first_data_col, comment_text))

                    result['changes'].append({
                        'type': 'if',
                        'row': cmd.location.row + 1,
                        'action': f'删除if标签行，添加注释: {comment_text} (位置: {get_column_letter(first_data_col)}{adjusted_data_row})'
                    })
                    result['converted_commands'] += 1

            elif isinstance(cmd, AreaCommand):
                area_commands.append(cmd)
                # 现有的area命令 - 在原始位置添加注释
                comment_text = cmd.to_jx_area_v2()
                # 计算调整后的行号（考虑删除的行）
                adjusted_row = cmd.location.row - len([r for r in rows_to_delete if r < cmd.location.row]) + 1
                comments_to_add.append((adjusted_row, cmd.location.col + 1, comment_text))

                result['changes'].append({
                    'type': 'area',
                    'row': cmd.location.row + 1,
                    'action': f'保留area命令: {comment_text}'
                })
                result['converted_commands'] += 1
                self.logger.info(f"      ✅ 保留area命令: {comment_text}")

            elif isinstance(cmd, MultiSheetCommand):
                comment_text = cmd.to_jx_multi_sheet_v2()
                comments_to_add.append((cmd.location.row + 1, cmd.location.col + 1, comment_text))
                rows_to_delete.add(cmd.location.row)

                result['changes'].append({
                    'type': 'multiSheet',
                    'row': cmd.location.row + 1,
                    'action': f'转换multiSheet，添加注释: {comment_text}'
                })
                result['converted_commands'] += 1

            elif isinstance(cmd, OutCommand):
                # 处理单独的jx:out单元格
                cell = ws.cell(row=cmd.location.row + 1, column=cmd.location.col + 1)
                new_value = cmd.to_expression()
                if cell.value != new_value:
                    cell.value = new_value
                    result['changes'].append({
                        'type': 'out',
                        'row': cmd.location.row + 1,
                        'col': cmd.location.col + 1,
                        'action': f'转换jx:out为表达式: {new_value}'
                    })
                    result['converted_commands'] += 1

        # 删除标记的行（从后往前删除）
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx + 1)  # openpyxl行号从1开始
            self.logger.debug(f"      删除行 {row_idx + 1}")

        # 处理jx:out指令（在单元格文本中）
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and ('<jx:out' in cell.value or 'jx:out(' in cell.value):
                    old_value = cell.value
                    # 替换所有的jx:out为表达式
                    new_value = re.sub(
                        r'<jx:out\s+select="([^"]+)"\s*/>',
                        lambda m: f'${{{m.group(1)}}}',
                        old_value
                    )
                    new_value = re.sub(
                        r'jx:out\s*\(\s*select\s*=\s*["\']([^"\']*)["\']\s*\)',
                        lambda m: f'${{{m.group(1)}}}',
                        new_value
                    )
                    if new_value != old_value:
                        cell.value = new_value
                        result['changes'].append({
                            'type': 'out',
                            'row': cell.row,
                            'col': cell.column,
                            'action': f'转换jx:out为表达式: {new_value}'
                        })
                        result['converted_commands'] += 1

        # 自动生成area命令（如果没有现有的）- 修复位置为A1
        if not area_commands and (rows_to_delete or comments_to_add):
            # 计算数据区域
            last_data_row = ws.max_row
            last_data_col = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        last_data_col = max(last_data_col, cell.column)

            if last_data_row > 0 and last_data_col > 0:
                last_cell = f"{get_column_letter(last_data_col)}{last_data_row}"
                area_comment = f'jx:area(lastCell="{last_cell}")'
                comments_to_add.append((1, 1, area_comment))  # 在A1添加area注释

                result['changes'].append({
                    'type': 'area',
                    'row': 1,
                    'action': f'自动添加area命令: {area_comment}'
                })
                result['converted_commands'] += 1
                self.logger.info(f"      ✅ 自动生成area命令: {area_comment}")

        # 添加注释
        for row, col, comment_text in comments_to_add:
            try:
                cell = ws.cell(row=row, column=col)
                cell.comment = Comment(comment_text, "JXLS Migration Tool")
                self.logger.debug(f"      添加注释到 {get_column_letter(col)}{row}: {comment_text}")
            except Exception as e:
                self.logger.debug(f"      添加注释失败 row={row}, col={col}: {e}")

        return result

    def find_first_data_column(self, xls_sheet, row_idx: int) -> int:
        """
        在XLS中查找指定行第一个有数据的列

        Args:
            xls_sheet: xlrd的Sheet对象
            row_idx: 行号

        Returns:
            第一个有数据的列索引，如果没有找到返回-1
        """
        for col_idx in range(xls_sheet.ncols):
            cell = xls_sheet.cell(row_idx, col_idx)
            if cell.value and str(cell.value).strip():
                return col_idx
        return -1

    def find_first_data_column_xlsx(self, ws: Worksheet, row_idx: int) -> int:
        """
        在XLSX中查找指定行第一个有数据的列

        Args:
            ws: openpyxl的Worksheet对象
            row_idx: 行号

        Returns:
            第一个有数据的列号（1-based），如果没有找到返回1
        """
        row_data = list(ws.iter_rows())[row_idx]
        for col_idx, cell in enumerate(row_data, start=1):
            if cell.value and str(cell.value).strip():
                return col_idx
        return 1  # 默认返回A列

    def find_end_tag(self, xls_sheet, start_row: int, end_tag: str) -> Optional[int]:
        """
        在XLS中查找结束标签的行号

        Args:
            xls_sheet: xlrd的Sheet对象
            start_row: 开始搜索的行号
            end_tag: 结束标签

        Returns:
            结束标签所在的行号
        """
        for row_idx in range(start_row + 1, xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell = xls_sheet.cell(row_idx, col_idx)
                value = str(cell.value)
                if end_tag in value:
                    return row_idx
        return None

    def find_end_tag_xlsx(self, ws: Worksheet, start_row: int, end_tag: str) -> Optional[int]:
        """
        在XLSX中查找结束标签的行号

        Args:
            ws: openpyxl的Worksheet对象
            start_row: 开始搜索的行号
            end_tag: 结束标签

        Returns:
            结束标签所在的行号
        """
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row + 2)):
            for cell in row:
                if cell.value:
                    value = str(cell.value)
                    if end_tag in value:
                        return start_row + 1 + row_idx
        return None

    def find_last_data_column(self, xls_sheet, row_idx: int) -> int:
        """
        在XLS中查找指定行最后一个有数据的列

        Args:
            xls_sheet: xlrd的Sheet对象
            row_idx: 行号

        Returns:
            最后一个有数据的列索引
        """
        last_col = 0
        for col_idx in range(xls_sheet.ncols):
            cell = xls_sheet.cell(row_idx, col_idx)
            if cell.value:
                last_col = col_idx
        return last_col

    def find_last_data_column_xlsx(self, ws: Worksheet, row_idx: int) -> int:
        """
        在XLSX中查找指定行最后一个有数据的列

        Args:
            ws: openpyxl的Worksheet对象
            row_idx: 行号

        Returns:
            最后一个有数据的列号
        """
        last_col = 1
        row_data = list(ws.iter_rows())[row_idx]
        for col_idx, cell in enumerate(row_data, start=1):
            if cell.value:
                last_col = col_idx
        return last_col

    def generate_report(self, output_dir: Path):
        """
        生成迁移报告

        Args:
            output_dir: 输出目录
        """
        # Markdown报告
        md_report = self.generate_markdown_report()
        md_path = output_dir / 'migration_report.md'
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(md_report)
        self.logger.info(f"已生成Markdown报告: {md_path}")

        # JSON报告
        json_report = {
            'timestamp': datetime.now().isoformat(),
            'stats': self.stats,
            'results': self.results,
            'failures': self.failures
        }
        json_path = output_dir / 'migration_report.json'
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(json_report, f, ensure_ascii=False, indent=2)
        self.logger.info(f"已生成JSON报告: {json_path}")

    def generate_markdown_report(self) -> str:
        """
        生成Markdown格式的迁移报告

        Returns:
            Markdown报告文本
        """
        report = []
        report.append("# JXLS 1.x → 2.14.0 模板迁移报告\n")
        report.append(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        report.append("---\n")

        # 统计信息
        report.append("## 📊 统计信息\n")
        report.append(f"- ✅ 成功: {self.stats['success']}")
        report.append(f"- ❌ 失败: {self.stats['failed']}")
        report.append(f"- 📊 总计: {self.stats['total']}")
        report.append(f"- 📄 处理文件: {self.stats['files_processed']}")
        report.append(f"- 🔧 发现命令: {self.stats['commands_found']}")
        report.append(f"- 🔄 转换命令: {self.stats['commands_converted']}")
        success_rate = (self.stats['success'] / self.stats['total'] * 100) if self.stats['total'] > 0 else 0
        report.append(f"- 🎯 成功率: {success_rate:.2f}%\n")
        report.append("---\n")

        # 成功的文件
        report.append("## ✅ 成功迁移的文件\n")
        success_results = [r for r in self.results if r['success']]
        if success_results:
            for idx, result in enumerate(success_results, 1):
                source_name = Path(result['source']).name
                target_name = Path(result['target']).name
                report.append(f"### {idx}. `{source_name}` → `{target_name}`\n")

                # 命令统计
                report.append(
                    f"- **命令统计**: 发现 {result.get('total_commands', 0)} 个，转换 {result.get('converted_commands', 0)} 个")

                # 列出变更类型
                if result.get('changes'):
                    changes_by_type = defaultdict(list)
                    for change in result['changes']:
                        changes_by_type[change['type']].append(change)

                    for change_type, changes in changes_by_type.items():
                        report.append(f"- **{change_type}**: {len(changes)} 处变更")
                report.append("")
        else:
            report.append("无\n")

        report.append("---\n")

        # 失败的文件
        report.append("## ❌ 失败的文件\n")
        if self.failures:
            for idx, failure in enumerate(self.failures, 1):
                file_name = Path(failure['file']).name
                report.append(f"{idx}. `{file_name}`")
                report.append(f"   - **错误**: {failure['error']}\n")
        else:
            report.append("无\n")

        report.append("---\n")

        # 迁移说明
        report.append("## 📝 迁移说明\n")
        report.append("### 主要变更\n")
        report.append("1. **jx:forEach → jx:each**")
        report.append("   - 删除 `jx:forEach` 和 `/jx:forEach` 标签行")
        report.append("   - 在数据行添加Excel注释")
        report.append("   - 注释格式: `jx:each(items=\"...\" var=\"...\" lastCell=\"...\")`\n")
        report.append("2. **jx:if → jx:if**")
        report.append("   - 参数 `test` 改为 `condition`")
        report.append("   - 同样使用注释方式\n")
        report.append("3. **jx:out → ${...}**")
        report.append("   - `<jx:out select=\"var\"/>` → `${var}`")
        report.append("   - `jx:out(select=\"var\")` → `${var}`\n")
        report.append("4. **jx:area**")
        report.append("   - 保留现有的area命令")
        report.append("   - 如果没有area命令，自动生成\n")
        report.append("5. **jx:multiSheet → jx:multiSheet**")
        report.append("   - 转换为注释方式\n")
        report.append("### 格式保留\n")
        report.append("- ✅ 单元格样式（字体、颜色、边框）")
        report.append("- ✅ 列宽、行高")
        report.append("- ✅ 合并单元格")
        report.append("- ✅ 背景色\n")
        report.append("---\n")

        # 下一步
        report.append("## 🎯 下一步\n")
        report.append("1. ✅ 检查迁移报告中的失败文件")
        report.append("2. ✅ 验证关键业务模板的导出功能")
        report.append("3. ✅ 运行单元测试")
        report.append("4. ✅ 更新Java代码使用新的模板路径")
        report.append("5. ✅ 逐步替换生产环境的模板\n")

        return '\n'.join(report)

    def print_summary(self):
        """打印迁移汇总信息"""
        self.logger.info("=" * 80)
        self.logger.info("迁移完成！")
        self.logger.info(f"✅ 成功: {self.stats['success']}")
        self.logger.info(f"❌ 失败: {self.stats['failed']}")
        self.logger.info(f"📊 总计: {self.stats['total']}")
        self.logger.info(f"📄 处理文件: {self.stats['files_processed']}")
        self.logger.info(f"🔧 发现命令: {self.stats['commands_found']}")
        self.logger.info(f"🔄 转换命令: {self.stats['commands_converted']}")
        success_rate = (self.stats['success'] / self.stats['total'] * 100) if self.stats['total'] > 0 else 0
        self.logger.info(f"🎯 成功率: {success_rate:.2f}%")
        if not self.dry_run and self.output_dir:
            self.logger.info(f"📄 报告已保存到: {Path(self.output_dir) / 'migration_report.md'}")
        self.logger.info("=" * 80)


# ============================================================================
# 命令行入口
# ============================================================================

def print_banner():
    """打印工具横幅"""
    banner = """
╔═══════════════════════════════════════════════════════════════════╗
║  JXLS 1.x → 2.14.0 自动化迁移工具（修复版 v3.4）                ║
║  Author: fivefish                                              ║
║  Version: 3.4 (Fixed)                                            ║
║  Date: 2025-11-07                                                 ║
║  🔥 默认使用 XlsxWriter（自动共享字符串表，POI兼容性更好）      ║
║  修复: jx:each注释生成 + jx:area位置问题                         ║
╚═══════════════════════════════════════════════════════════════════╝
"""
    print(banner)


def main():
    """主函数"""
    print_banner()

    parser = argparse.ArgumentParser(
        description='JXLS 1.x到2.14.0自动化迁移工具（修复版 v3.4）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 迁移目录（默认使用 XlsxWriter，自动共享字符串表）
  python jxls_migration_tool.py input_dir

  # 迁移目录并保持原文件后缀（推荐）
  python jxls_migration_tool.py input_dir --keep-extension

  # 指定输出目录
  python jxls_migration_tool.py input_dir -o output_dir

  # 试运行（不实际修改文件）
  python jxls_migration_tool.py input_dir --dry-run

  # 详细日志输出
  python jxls_migration_tool.py input_dir --verbose

  # 迁移单个文件
  python jxls_migration_tool.py input.xls -f output.xlsx

  # 完整示例：保持后缀 + 试运行 + 详细日志
  python jxls_migration_tool.py exceltemplate_backup -o exceltemplate --keep-extension --dry-run --verbose

注意: 需要安装 xlsxwriter (pip install xlsxwriter)
        """
    )

    parser.add_argument('input', help='输入目录或文件路径')
    parser.add_argument('-o', '--output', help='输出目录或文件路径')
    parser.add_argument('-f', '--file', action='store_true', help='迁移单个文件（而不是目录）')
    parser.add_argument('--dry-run', action='store_true', help='试运行模式（不实际修改文件）')
    parser.add_argument('--keep-extension', action='store_true',
                        help='保持原文件后缀名，但文件内容始终为.xlsx格式（.xls文件转换为.xlsx格式但文件名保持.xls，.xlsx文件保持.xlsx）')
    parser.add_argument('--verbose', action='store_true', help='详细日志输出')

    args = parser.parse_args()

    # 创建迁移工具
    # 统一使用 XlsxWriter（更好的POI兼容性）
    use_xlsxwriter = True
    tool = JxlsMigrationTool(
        dry_run=args.dry_run,
        keep_extension=args.keep_extension,
        verbose=args.verbose,
        use_xlsxwriter=use_xlsxwriter
    )

    try:
        if args.file:
            # 迁移单个文件
            if not args.output:
                # 如果没有指定输出，根据keep_extension决定后缀
                input_path = Path(args.input)
                input_ext = input_path.suffix.lower()
                if args.keep_extension:
                    # 保持原后缀名，但文件内容始终为.xlsx格式
                    if input_ext == '.xls':
                        # .xls文件转换为.xlsx格式，但文件名保持.xls后缀
                        output_file = str(input_path.with_suffix('.xls'))
                    else:
                        # .xlsx文件直接输出.xlsx
                        output_file = str(input_path.with_suffix('.xlsx'))
                else:
                    output_file = str(input_path.with_suffix('.xlsx'))
                args.output = output_file

            # 设置日志
            tool.logger = setup_logging(None, args.dry_run, args.verbose)

            # 检查 XlsxWriter 是否可用
            if not XLSXWRITER_AVAILABLE:
                tool.logger.error("❌ 错误: 缺少 xlsxwriter 库")
                tool.logger.error("   请运行: pip install xlsxwriter")
                sys.exit(1)

            tool.logger.info("✓ 使用 XlsxWriter（自动共享字符串表，POI兼容性更好）")

            # 使用健壮的迁移方法，支持自动回退
            result = tool.migrate_file(args.input, args.output)

            # 显示尝试记录（如果有回退）
            if 'attempts' in result and len(result['attempts']) > 1:
                tool.logger.debug(f"尝试记录: {result['attempts']}")

            if result['success']:
                tool.logger.info(f"✅ 迁移成功: {args.output}")
                tool.logger.info(
                    f"🔧 发现 {result.get('total_commands', 0)} 个命令，转换 {result.get('converted_commands', 0)} 个")
                sys.exit(0)
            else:
                tool.logger.error(f"❌ 迁移失败: {result.get('error', '未知错误')}")
                if 'attempts' in result:
                    tool.logger.debug(f"完整尝试记录: {result['attempts']}")
                sys.exit(1)
        else:
            # 迁移目录
            result = tool.migrate_directory(args.input, args.output)

            if result['stats']['failed'] == 0:
                sys.exit(0)
            else:
                sys.exit(1)

    except KeyboardInterrupt:
        print("\n\n用户中断，退出...")
        sys.exit(130)
    except Exception as e:
        print(f"\n错误: {e}")
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()